import csv
import datetime
import os
import sqlite3

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

IMG_DIR = "D:\Apps\\umas\\export\\Texture2D"
OUTPUT_XLSX = "results/gacha_with_image.xlsx"
OUTPUT_CSV = "results/gacha_summary.csv"


def ts2str(ts):
    return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


conn = sqlite3.connect("master.mdb")
cur = conn.cursor()

cur.execute("""
    SELECT 
        ga.gacha_id, 
        ga.card_id, 
        ga.card_type, 
        gd.start_date, 
        gd.end_date, 
        td.text
    FROM gacha_available ga
    JOIN gacha_data gd 
        ON ga.gacha_id = gd.id
    JOIN text_data td 
        ON ga.card_id = td."index"
       AND td.category = CASE ga.card_type 
                            WHEN 1 THEN 4
                            WHEN 2 THEN 75
                         END
    WHERE ga.is_pickup = 1
""")

rows = cur.fetchall()
conn.close()

# 2. 聚合卡池
grouped = {}
for gacha_id, card_id, card_type, start, end, name in rows:
    pool_type = "角色卡池" if card_type == 1 else "支援卡"
    key = (gacha_id, pool_type)
    if key not in grouped:
        grouped[key] = {
            "卡池ID": gacha_id,
            "卡池类型": pool_type,
            "PickUp卡名": [name],
            "开始时间": ts2str(start),
            "结束时间": ts2str(end),
            "开始时间戳": start,
        }
    else:
        grouped[key]["PickUp卡名"].append(name)

# 3. 转成列表并聚合名字 + 图片路径
result = []
for g in grouped.values():
    g["PickUp卡名"] = "、".join(g["PickUp卡名"])
    img_file = f"img_bnr_gacha_{g['卡池ID']}.png"
    if os.path.exists(os.path.join(IMG_DIR, img_file)):
        g["图片"] = os.path.join(IMG_DIR, img_file)
    else:
        g["图片"] = None
    result.append(g)

# 4. 排序
result.sort(key=lambda r: (r["开始时间戳"], 0 if r["卡池类型"] == "支援卡" else 1))

for output_path in (OUTPUT_XLSX, OUTPUT_CSV):
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as csv_file:
    writer = csv.writer(csv_file)
    writer.writerow(
        ["卡池ID", "卡池类型", "开始时间", "结束时间", "PickUp卡名", "图片路径"]
    )
    for r in result:
        writer.writerow(
            [
                r["卡池ID"],
                r["卡池类型"],
                r["开始时间"],
                r["结束时间"],
                r["PickUp卡名"],
                r.get("图片") or "",
            ]
        )

wb = Workbook()
ws = wb.active
ws.title = "卡池列表"

# 表头
headers = ["日期", "类型", "PickUp卡", "图片"]
ws.append(headers)

# 设置表头样式
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(name="微软雅黑", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 交替背景色
colors = ["FFDCE6F1", "FFFFF2CC"]  # 浅蓝和浅黄
last_start_time = None
color_idx = 0

for r_idx, r in enumerate(result, start=2):
    # 交替颜色
    if r["开始时间"] != last_start_time:
        color_idx = 1 - color_idx
    last_start_time = r["开始时间"]
    fill = PatternFill(
        start_color=colors[color_idx], end_color=colors[color_idx], fill_type="solid"
    )

    # 卡池类型
    ws.cell(row=r_idx, column=2, value=r["卡池类型"]).fill = fill
    ws.cell(row=r_idx, column=2).font = Font(name="微软雅黑", bold=True)
    ws.cell(row=r_idx, column=2).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    # PickUp卡名（C列自动换行）
    ws.cell(row=r_idx, column=3, value=r["PickUp卡名"]).fill = fill
    ws.cell(row=r_idx, column=3).font = Font(name="微软雅黑", bold=True)
    ws.cell(row=r_idx, column=3).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 日期合并列（D列显示“开始~结束”）
    ws.cell(row=r_idx, column=1, value=f"{r['开始时间']}\n~{r['结束时间']}").fill = fill
    ws.cell(row=r_idx, column=1).font = Font(name="微软雅黑", bold=True)
    ws.cell(row=r_idx, column=1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.cell(row=r_idx, column=4).fill = fill

    # 插入图片
    if r.get("图片") and os.path.exists(r["图片"]):
        img = XLImage(r["图片"])
        # 控制宽度
        scale_factor = 500 / img.width
        img.width = int(img.width * scale_factor)
        img.height = int(img.height * scale_factor)
        ws.add_image(img, f"D{r_idx}")

        # 行高自适应（图片或文字换行）
    ws.row_dimensions[r_idx].height = 100


# --- 合并日期列 D ---
merge_start = 2
last_start_time = None
color_idx = 0
last_color = 0

for r_idx, r in enumerate(result, start=2):
    if r["开始时间"] != last_start_time:
        color_idx = 1 - color_idx
        last_start_time = r["开始时间"]
    if color_idx != last_color:
        # 上一段结束，合并
        if merge_start < r_idx - 1:
            ws.merge_cells(
                start_row=merge_start,
                start_column=1,
                end_row=r_idx - 1,
                end_column=1,
            )
            ws.cell(
                row=merge_start,
                column=1,
                value=f"{result[merge_start - 2]['开始时间']}\n~{result[r_idx - 3]['结束时间']}",
            )
            ws.cell(row=merge_start, column=1).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        merge_start = r_idx
        last_color = color_idx  # type:ignore


if merge_start < len(result) + 1:
    ws.merge_cells(
        start_row=merge_start, start_column=4, end_row=len(result) + 1, end_column=4
    )
    ws.cell(
        row=merge_start,
        column=1,
        value=f"{result[merge_start - 2]['开始时间']}\n~{result[-1]['结束时间']}",
    )
    ws.cell(row=merge_start, column=1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )


# 列宽调整
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 70

wb.save(OUTPUT_XLSX)
print(f"✅ 已生成简易 CSV: {OUTPUT_CSV}")
print(f"✅ 已生成带颜色和合并单元格的 Excel: {OUTPUT_XLSX}")
